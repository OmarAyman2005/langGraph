import sys
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))
sys.path.insert(0, str(PROJECT_ROOT / "tests"))

from normalizer.normalizer import normalize_raw_prompt
from parsers.normalized_prompt_parser import parse_normalized_prompt
from final_run_utils import calculate_actual_normalization_counts


# ============================================================
# CONFIG
# ============================================================

DATASET_FILE = Path(r"C:\Users\Omar Ayman\Desktop\Handassa\SEM8 (BACHELOR)\Testing Excel Files\DataSet.xlsx")
REPORT_FILE = Path(r"C:\Users\Omar Ayman\Desktop\Handassa\SEM8 (BACHELOR)\Testing Excel Files\Dataset_Validation_Report.xlsx")


D1_REQUIRED_COLUMNS = [
    "Example_ID",
    "Dataset_Type",
    "Symbolic_Skeleton",
    "Clean_Raw_Input",
    "Raw_Input",
    "Entailment_Status",
    "Not_Entailed_Type",
    "Inference_Depth",
    "Inference_Rules",
    "Distractor_Count",
    "Case_Adjustment_Count",
    "Pattern_Rewrite_Count",
    "Subject_Propagation_Count",
    "Synonym_Unification_Count",
    "Antonym_Unification_Count",
    "Normalization_Complexity_Score",
]


D2_REQUIRED_COLUMNS = [
    "Example_ID",
    "Dataset_Type",
    "Symbolic_Skeleton",
    "Clean_Raw_Input",
    "Raw_Input",
    "Expected_Component",
    "Expected_SubComponent",
    "Expected_Specific_Error",
]


COUNT_COLUMNS = [
    "Case_Adjustment_Count",
    "Pattern_Rewrite_Count",
    "Subject_Propagation_Count",
    "Synonym_Unification_Count",
    "Antonym_Unification_Count",
]


# ============================================================
# GENERAL HELPERS
# ============================================================

def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_label(value: Any) -> str:
    return clean(value).lower()


def to_int(value: Any, default: Optional[int] = None) -> Optional[int]:
    try:
        if value is None or clean(value) == "":
            return default
        return int(float(value))
    except Exception:
        return default


def get_header_map(ws) -> Dict[str, int]:
    header_map = {}

    for col_idx, cell in enumerate(ws[1], start=1):
        name = clean(cell.value)

        if name:
            header_map[name] = col_idx

    return header_map


def row_to_dict(ws, row_idx: int, header_map: Dict[str, int]) -> Dict[str, Any]:
    data = {}

    for header, col_idx in header_map.items():
        data[header] = ws.cell(row=row_idx, column=col_idx).value

    return data


def has_required_headers(header_map: Dict[str, int], required: List[str]) -> bool:
    return all(col in header_map for col in required)


def find_dataset_sheets(wb) -> List[str]:
    """
    Finds sheets that look like dataset sheets by checking for Example_ID and Raw_Input.
    """

    dataset_sheets = []

    for ws in wb.worksheets:
        header_map = get_header_map(ws)

        if "Example_ID" in header_map and "Raw_Input" in header_map:
            dataset_sheets.append(ws.title)

    return dataset_sheets


def add_result(
    results: List[Dict[str, Any]],
    example_id: str,
    dataset_type: str,
    sheet_name: str,
    check_category: str,
    field: str,
    expected: Any,
    actual: Any,
    status: str,
    notes: str = "",
) -> None:
    results.append(
        {
            "Example_ID": example_id,
            "Dataset_Type": dataset_type,
            "Sheet": sheet_name,
            "Check_Category": check_category,
            "Field": field,
            "Expected": expected,
            "Actual": actual,
            "Status": status,
            "Notes": notes,
        }
    )


# ============================================================
# VALIDATION CHECKS
# ============================================================

def validate_required_fields(
    row: Dict[str, Any],
    sheet_name: str,
    results: List[Dict[str, Any]],
) -> None:
    example_id = clean(row.get("Example_ID"))
    dataset_type = clean(row.get("Dataset_Type"))

    if dataset_type == "D1":
        required = D1_REQUIRED_COLUMNS
    elif dataset_type == "D2":
        required = D2_REQUIRED_COLUMNS
    else:
        add_result(
            results,
            example_id,
            dataset_type,
            sheet_name,
            "Required Fields",
            "Dataset_Type",
            "D1 or D2",
            dataset_type,
            "FAIL",
            "Unknown Dataset_Type.",
        )
        return

    expected_error = clean(row.get("Expected_Specific_Error"))

    for field in required:
        value = row.get(field)

        # D2_EX001 intentionally tests empty raw input.
        # Therefore Raw_Input and Clean_Raw_Input are allowed to be empty
        # when the expected error is CA_EMPTY_INPUT.
        if (
            dataset_type == "D2"
            and expected_error == "CA_EMPTY_INPUT"
            and field in {"Raw_Input", "Clean_Raw_Input"}
        ):
            add_result(
                results,
                example_id,
                dataset_type,
                sheet_name,
                "Required Fields",
                field,
                "Allowed empty for CA_EMPTY_INPUT",
                value,
                "PASS",
                "This D2 row intentionally tests empty input.",
            )
            continue

        if value is None or clean(value) == "":
            add_result(
                results,
                example_id,
                dataset_type,
                sheet_name,
                "Required Fields",
                field,
                "Non-empty",
                value,
                "FAIL",
                "Required field is empty.",
            )


def validate_complexity_sum(
    row: Dict[str, Any],
    sheet_name: str,
    results: List[Dict[str, Any]],
) -> None:
    example_id = clean(row.get("Example_ID"))
    dataset_type = clean(row.get("Dataset_Type"))

    if dataset_type != "D1":
        return

    expected_counts = {}

    for col in COUNT_COLUMNS:
        expected_counts[col] = to_int(row.get(col), default=None)

    expected_score = to_int(row.get("Normalization_Complexity_Score"), default=None)

    if any(value is None for value in expected_counts.values()) or expected_score is None:
        add_result(
            results,
            example_id,
            dataset_type,
            sheet_name,
            "Complexity Sum",
            "Normalization_Complexity_Score",
            "Valid integer sum",
            "Missing or non-integer value",
            "FAIL",
            "One or more normalization count columns are invalid.",
        )
        return

    actual_score = sum(expected_counts.values())

    status = "PASS" if actual_score == expected_score else "FAIL"

    add_result(
        results,
        example_id,
        dataset_type,
        sheet_name,
        "Complexity Sum",
        "Normalization_Complexity_Score",
        expected_score,
        actual_score,
        status,
        "Checks whether the score equals the sum of the five normalization count fields.",
    )


def validate_d1_metadata_logic_basic(
    row: Dict[str, Any],
    sheet_name: str,
    results: List[Dict[str, Any]],
) -> None:
    """
    Basic deterministic metadata checks.
    This does NOT prove full logical correctness yet.
    It catches obvious metadata mistakes.
    """

    example_id = clean(row.get("Example_ID"))
    dataset_type = clean(row.get("Dataset_Type"))

    if dataset_type != "D1":
        return

    entailment_status = normalize_label(row.get("Entailment_Status"))
    ne_type = clean(row.get("Not_Entailed_Type"))

    if entailment_status == "entailed":
        status = "PASS" if ne_type == "N/A" else "FAIL"
        add_result(
            results,
            example_id,
            dataset_type,
            sheet_name,
            "Basic D1 Metadata",
            "Not_Entailed_Type",
            "N/A when Entailment_Status = entailed",
            ne_type,
            status,
            "Entailed examples must not have NE1/NE2/NE3.",
        )

    elif entailment_status == "not_entailed":
        status = "PASS" if ne_type in {"NE1", "NE2", "NE3"} else "FAIL"
        add_result(
            results,
            example_id,
            dataset_type,
            sheet_name,
            "Basic D1 Metadata",
            "Not_Entailed_Type",
            "NE1/NE2/NE3 when Entailment_Status = not_entailed",
            ne_type,
            status,
            "Not-entailed examples must specify a not-entailment type.",
        )

    else:
        add_result(
            results,
            example_id,
            dataset_type,
            sheet_name,
            "Basic D1 Metadata",
            "Entailment_Status",
            "entailed or not_entailed",
            row.get("Entailment_Status"),
            "FAIL",
            "Invalid entailment label.",
        )


def validate_d1_normalizer_counts(
    row: Dict[str, Any],
    sheet_name: str,
    results: List[Dict[str, Any]],
) -> None:
    example_id = clean(row.get("Example_ID"))
    dataset_type = clean(row.get("Dataset_Type"))

    if dataset_type != "D1":
        return

    raw_input = clean(row.get("Raw_Input"))

    if not raw_input:
        add_result(
            results,
            example_id,
            dataset_type,
            sheet_name,
            "Normalizer Counts",
            "Raw_Input",
            "Non-empty raw input",
            raw_input,
            "FAIL",
            "Cannot run normalizer on empty Raw_Input.",
        )
        return

    try:
        normalizer_result = normalize_raw_prompt(raw_input)
    except Exception as exc:
        add_result(
            results,
            example_id,
            dataset_type,
            sheet_name,
            "Normalizer Execution",
            "Normalizer_Status",
            "success",
            f"exception: {exc}",
            "FAIL",
            "Normalizer crashed.",
        )
        return

    if normalizer_result.get("success") is not True:
        add_result(
            results,
            example_id,
            dataset_type,
            sheet_name,
            "Normalizer Execution",
            "Normalizer_Status",
            "success",
            "failed",
            "FAIL",
            normalizer_result.get("error", "Unknown normalizer error."),
        )
        return

    actual_counts = calculate_actual_normalization_counts(raw_input, normalizer_result)

    comparisons = [
        ("Case_Adjustment_Count", actual_counts["case"]),
        ("Synonym_Unification_Count", actual_counts["synonym"]),
        ("Antonym_Unification_Count", actual_counts["antonym"]),
    ]

    # N4 and N6 are currently manually reviewed because the current normalizer
    # output/debug structure does not expose reliable actual rewrite/propagation
    # counts to this validator.
    for field in ["Pattern_Rewrite_Count", "Subject_Propagation_Count"]:
        add_result(
            results,
            example_id,
            dataset_type,
            sheet_name,
            "Normalizer Counts",
            field,
            row.get(field),
            "Manual review required",
            "REVIEW",
            "Current validator cannot reliably extract this count from normalizer output.",
        )

    # Recalculate partial actual score only for automatically reliable fields.
    expected_total = to_int(row.get("Normalization_Complexity_Score"), default=None)
    manual_pattern = to_int(row.get("Pattern_Rewrite_Count"), default=0)
    manual_subject = to_int(row.get("Subject_Propagation_Count"), default=0)

    actual_total_with_manual_review = (
        actual_counts["case"]
        + manual_pattern
        + manual_subject
        + actual_counts["synonym"]
        + actual_counts["antonym"]
    )

    status = "PASS" if expected_total == actual_total_with_manual_review else "FAIL"

    add_result(
        results,
        example_id,
        dataset_type,
        sheet_name,
        "Normalizer Counts",
        "Normalization_Complexity_Score",
        expected_total,
        actual_total_with_manual_review,
        status,
        "Total score checked using automatic case/synonym/antonym counts and manually reviewed pattern/subject counts.",
    )

    for field, actual in comparisons:
        expected = to_int(row.get(field), default=None)

        status = "PASS" if expected == actual else "FAIL"

        add_result(
            results,
            example_id,
            dataset_type,
            sheet_name,
            "Normalizer Counts",
            field,
            expected,
            actual,
            status,
            "Expected value from dataset compared against actual normalizer-triggered value.",
        )


def validate_d1_parser_safety(
    row: Dict[str, Any],
    sheet_name: str,
    results: List[Dict[str, Any]],
) -> None:
    example_id = clean(row.get("Example_ID"))
    dataset_type = clean(row.get("Dataset_Type"))

    if dataset_type != "D1":
        return

    raw_input = clean(row.get("Raw_Input"))

    if not raw_input:
        return

    try:
        normalizer_result = normalize_raw_prompt(raw_input)

        if normalizer_result.get("success") is not True:
            return

        normalized_prompt = normalizer_result.get("normalized_input")
        parser_result = parse_normalized_prompt(normalized_prompt)

        if parser_result.get("prompt_parse_success") is True:
            status = "PASS"
            actual = "success"
            notes = "Raw_Input is safe through Normalizer and Prompt Parser."
        else:
            status = "FAIL"
            actual = "failed"
            notes = parser_result.get("prompt_parse_error", "Unknown parser error.")

        add_result(
            results,
            example_id,
            dataset_type,
            sheet_name,
            "Pipeline Safety",
            "Prompt_Parser_Status",
            "success",
            actual,
            status,
            notes,
        )

    except Exception as exc:
        add_result(
            results,
            example_id,
            dataset_type,
            sheet_name,
            "Pipeline Safety",
            "Prompt_Parser_Status",
            "success",
            f"exception: {exc}",
            "FAIL",
            "Exception while checking parser safety.",
        )


def add_logic_review_marker(
    row: Dict[str, Any],
    sheet_name: str,
    results: List[Dict[str, Any]],
) -> None:
    """
    Important:
    This script cannot fully prove Entailment_Status, Inference_Depth, and Inference_Rules
    unless Symbolic_Skeleton is written in a strict machine-readable proof format.

    So for now, it marks logical checks as requiring review.
    """

    example_id = clean(row.get("Example_ID"))
    dataset_type = clean(row.get("Dataset_Type"))

    if dataset_type != "D1":
        return

    for field in [
        "Entailment_Status",
        "Not_Entailed_Type",
        "Inference_Depth",
        "Inference_Rules",
        "Distractor_Count",
    ]:
        add_result(
            results,
            example_id,
            dataset_type,
            sheet_name,
            "Logical Proof Validation",
            field,
            row.get(field),
            "Needs strict symbolic proof validation",
            "REVIEW",
            "Automatic proof validation requires Symbolic_Skeleton to use a strict parseable proof format.",
        )


def validate_row(
    row: Dict[str, Any],
    sheet_name: str,
    results: List[Dict[str, Any]],
) -> None:
    validate_required_fields(row, sheet_name, results)
    validate_complexity_sum(row, sheet_name, results)
    validate_d1_metadata_logic_basic(row, sheet_name, results)
    validate_d1_normalizer_counts(row, sheet_name, results)
    validate_d1_parser_safety(row, sheet_name, results)
    add_logic_review_marker(row, sheet_name, results)


# ============================================================
# REPORT WRITING
# ============================================================

def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="000000")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

            if cell.column_letter == "H":
                value = clean(cell.value)

                if value == "PASS":
                    cell.fill = PatternFill("solid", fgColor="C6EFCE")
                elif value == "FAIL":
                    cell.fill = PatternFill("solid", fgColor="FFC7CE")
                elif value == "REVIEW":
                    cell.fill = PatternFill("solid", fgColor="FFEB9C")

    widths = {
        "A": 14,
        "B": 12,
        "C": 24,
        "D": 24,
        "E": 34,
        "F": 28,
        "G": 28,
        "H": 12,
        "I": 60,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_report(results: List[Dict[str, Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Validation_Report"

    headers = [
        "Example_ID",
        "Dataset_Type",
        "Sheet",
        "Check_Category",
        "Field",
        "Expected",
        "Actual",
        "Status",
        "Notes",
    ]

    ws.append(headers)

    for result in results:
        ws.append([result.get(header, "") for header in headers])

    style_sheet(ws)

    summary = wb.create_sheet("Summary")
    summary_headers = ["Status", "Count"]
    summary.append(summary_headers)

    statuses = ["PASS", "FAIL", "REVIEW"]
    for status in statuses:
        count = sum(1 for result in results if result["Status"] == status)
        summary.append([status, count])

    style_sheet(summary)

    wb.save(REPORT_FILE)


# ============================================================
# MAIN
# ============================================================

def main() -> None:
    if not DATASET_FILE.exists():
        raise FileNotFoundError(f"Dataset file not found: {DATASET_FILE}")

    wb = load_workbook(DATASET_FILE, data_only=False)
    dataset_sheets = find_dataset_sheets(wb)

    if not dataset_sheets:
        raise RuntimeError("No dataset sheets found. Expected sheets with Example_ID and Raw_Input headers.")

    results = []

    for sheet_name in dataset_sheets:
        ws = wb[sheet_name]
        header_map = get_header_map(ws)

        for row_idx in range(2, ws.max_row + 1):
            row = row_to_dict(ws, row_idx, header_map)
            example_id = clean(row.get("Example_ID"))

            if not example_id:
                continue

            validate_row(row, sheet_name, results)

    write_report(results)

    fail_count = sum(1 for result in results if result["Status"] == "FAIL")
    review_count = sum(1 for result in results if result["Status"] == "REVIEW")
    pass_count = sum(1 for result in results if result["Status"] == "PASS")

    print("=" * 100)
    print("DATASET VALIDATION FINISHED")
    print("=" * 100)
    print(f"Dataset file: {DATASET_FILE}")
    print(f"Report file:  {REPORT_FILE}")
    print()
    print(f"PASS:   {pass_count}")
    print(f"FAIL:   {fail_count}")
    print(f"REVIEW: {review_count}")
    print("=" * 100)


if __name__ == "__main__":
    main()